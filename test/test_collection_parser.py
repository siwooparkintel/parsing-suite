"""
Integration tests for Collection_Parser.py.

Each test:
- Generates a collection JSON dynamically with absolute paths to fixture CSVs
- Runs Collection_Parser.py as a subprocess
- Checks exit code 0
- Verifies the output _collection.xlsx exists with expected structure
"""
from __future__ import annotations

import json
import openpyxl
import pytest
from pathlib import Path

from conftest import run_parser, TEST_CONFIG, FIXTURE_COLLECTION_PWR


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _build_collection_json(entries: list[dict], dest: Path) -> Path:
    """Write collection JSON to dest and return its path."""
    dest.write_text(json.dumps(entries, indent=2), encoding="utf-8")
    return dest


def _run_collection_parser(tmp_path: Path, collection_json: Path, config: Path = TEST_CONFIG):
    out_prefix = str(tmp_path / "result")
    result = run_parser(
        "Collection_Parser.py",
        ["-c", str(config), "-i", str(collection_json), "-o", out_prefix],
    )
    return result, Path(out_prefix + "_collection.xlsx")


def _baseline_csv() -> str:
    return str((FIXTURE_COLLECTION_PWR / "run_baseline" / "pacs-summary.csv").resolve())


def _optimized_csv() -> str:
    return str((FIXTURE_COLLECTION_PWR / "run_optimized" / "pacs-summary.csv").resolve())


def _two_entry_collection() -> list[dict]:
    return [
        {
            "data_label": "Baseline",
            "condition": "baseline_run",
            "data_summary_type": "compact",
            "power_summary_path": _baseline_csv(),
        },
        {
            "data_label": "Optimized",
            "condition": "optimized_run",
            "data_summary_type": "compact",
            "power_summary_path": _optimized_csv(),
        },
    ]


def _load_attribute_column(xlsx_path: Path) -> list[str]:
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    values = [row[0].value for row in ws.iter_rows() if row[0].value is not None]
    wb.close()
    return values


# ---------------------------------------------------------------------------
# tests
# ---------------------------------------------------------------------------

def test_help_flag():
    """Collection_Parser -h exits 0 and shows usage information."""
    result = run_parser("Collection_Parser.py", ["-h"])
    assert result.returncode == 0
    output = (result.stdout or "") + (result.stderr or "")
    assert "Collection Parser" in output or "usage" in output.lower()


def test_power_only_collection_exits_zero(tmp_path: Path):
    """Collection_Parser exits 0 for a minimal power-only collection."""
    cj = _build_collection_json(_two_entry_collection(), tmp_path / "collection.json")
    result, xlsx = _run_collection_parser(tmp_path, cj)
    assert result.returncode == 0, f"Collection_Parser failed:\n{result.stderr}"


def test_output_excel_created(tmp_path: Path):
    """Output _collection.xlsx must be created."""
    cj = _build_collection_json(_two_entry_collection(), tmp_path / "collection.json")
    _, xlsx = _run_collection_parser(tmp_path, cj)
    assert xlsx.exists(), f"Expected output not found: {xlsx}"


def test_output_has_attribute_column(tmp_path: Path):
    """Transposed output must have 'Attribute' as the first column header."""
    cj = _build_collection_json(_two_entry_collection(), tmp_path / "collection.json")
    _, xlsx = _run_collection_parser(tmp_path, cj)
    wb = openpyxl.load_workbook(str(xlsx), read_only=True)
    ws = wb.active
    first_header = ws.cell(row=1, column=1).value
    wb.close()
    assert first_header == "Attribute", f"Expected 'Attribute' header, got: {first_header!r}"


def test_output_has_data_label_row(tmp_path: Path):
    """Attribute column must contain a 'Data_label' row (Collection_Parser uses underscore)."""
    cj = _build_collection_json(_two_entry_collection(), tmp_path / "collection.json")
    _, xlsx = _run_collection_parser(tmp_path, cj)
    attributes = _load_attribute_column(xlsx)
    assert "Data_label" in attributes, f"'Data_label' row missing. Found: {attributes}"


def test_output_has_power_rail_attributes(tmp_path: Path):
    """P_SOC and P_VCCCORE power rails must appear as attribute rows in the output."""
    cj = _build_collection_json(_two_entry_collection(), tmp_path / "collection.json")
    _, xlsx = _run_collection_parser(tmp_path, cj)
    attributes = _load_attribute_column(xlsx)
    assert "P_SOC" in attributes, f"'P_SOC' attribute missing. Found: {attributes}"
    assert "P_VCCCORE" in attributes, f"'P_VCCCORE' attribute missing. Found: {attributes}"


def test_two_entries_produce_two_data_columns(tmp_path: Path):
    """Two collection entries must produce exactly 2 data columns (+1 Attribute col = 3 total)."""
    cj = _build_collection_json(_two_entry_collection(), tmp_path / "collection.json")
    _, xlsx = _run_collection_parser(tmp_path, cj)
    wb = openpyxl.load_workbook(str(xlsx), read_only=True)
    ws = wb.active
    max_col = ws.max_column
    wb.close()
    assert max_col == 3, f"Expected 3 columns (Attribute + 2 datasets), got {max_col}"


def test_both_labels_appear_in_output(tmp_path: Path):
    """The 'Data_label' row values must include 'Baseline' and 'Optimized'."""
    cj = _build_collection_json(_two_entry_collection(), tmp_path / "collection.json")
    _, xlsx = _run_collection_parser(tmp_path, cj)
    wb = openpyxl.load_workbook(str(xlsx), read_only=True)
    ws = wb.active

    # Find the row where column A == "Data_label"
    label_row_values = []
    for row in ws.iter_rows():
        if row[0].value == "Data_label":
            label_row_values = [cell.value for cell in row[1:]]
            break
    wb.close()

    assert "Baseline" in label_row_values, (
        f"'Baseline' not found in Data_label row. Values: {label_row_values}"
    )
    assert "Optimized" in label_row_values, (
        f"'Optimized' not found in Data_label row. Values: {label_row_values}"
    )


def test_single_entry_collection(tmp_path: Path):
    """A single-entry collection should still produce a valid Excel with 2 columns."""
    entries = [
        {
            "data_label": "OnlyRun",
            "condition": "only_run",
            "data_summary_type": "expanded",
            "power_summary_path": _baseline_csv(),
        }
    ]
    cj = _build_collection_json(entries, tmp_path / "collection.json")
    result, xlsx = _run_collection_parser(tmp_path, cj)
    assert result.returncode == 0, f"Collection_Parser failed:\n{result.stderr}"
    assert xlsx.exists()
    wb = openpyxl.load_workbook(str(xlsx), read_only=True)
    ws = wb.active
    max_col = ws.max_column
    wb.close()
    assert max_col == 2, f"Expected 2 columns (Attribute + 1 dataset), got {max_col}"
