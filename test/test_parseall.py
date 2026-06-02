"""
Integration tests for ParseAll.py.

Each test runs ParseAll.py as a subprocess, checks exit code, verifies the
output Excel file exists and contains expected structural content.
"""
from __future__ import annotations

import openpyxl
import pytest
from pathlib import Path

from conftest import run_parser, TEST_CONFIG, FIXTURE_PARSEALL_INPUT


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _run_parseall(tmp_path: Path, input_dir: Path, config: Path = TEST_CONFIG):
    """Run ParseAll with the given input folder and a temp output prefix."""
    out_prefix = str(tmp_path / "result")
    result = run_parser(
        "ParseAll.py",
        ["-c", str(config), "-i", str(input_dir), "-o", out_prefix],
    )
    return result, Path(out_prefix + "_allPower_v.xlsx")


def _load_attribute_column(xlsx_path: Path) -> list[str]:
    """Return all values in the first column ('Attribute') of the v-excel."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    values = [row[0].value for row in ws.iter_rows() if row[0].value is not None]
    wb.close()
    return values


# ---------------------------------------------------------------------------
# tests
# ---------------------------------------------------------------------------

def test_help_flag():
    """ParseAll -h exits 0 and shows usage information."""
    result = run_parser("ParseAll.py", ["-h"])
    assert result.returncode == 0
    output = (result.stdout or "") + (result.stderr or "")
    assert "AI summary parser" in output or "usage" in output.lower()


def test_power_only_run(tmp_path: Path):
    """ParseAll exits 0 and creates _allPower_v.xlsx for a power-only folder."""
    result, xlsx = _run_parseall(tmp_path, FIXTURE_PARSEALL_INPUT)
    assert result.returncode == 0, f"ParseAll failed:\n{result.stderr}"
    assert xlsx.exists(), f"Expected output not found: {xlsx}"


def test_output_excel_has_attribute_column(tmp_path: Path):
    """Transposed output must have 'Attribute' as the first column header."""
    _, xlsx = _run_parseall(tmp_path, FIXTURE_PARSEALL_INPUT)
    wb = openpyxl.load_workbook(str(xlsx), read_only=True)
    ws = wb.active
    first_header = ws.cell(row=1, column=1).value
    wb.close()
    assert first_header == "Attribute", f"Expected 'Attribute' header, got: {first_header!r}"


def test_output_has_data_label_row(tmp_path: Path):
    """Attribute column must contain a 'Data label' row (ParseAll uses list labels)."""
    _, xlsx = _run_parseall(tmp_path, FIXTURE_PARSEALL_INPUT)
    attributes = _load_attribute_column(xlsx)
    assert "Data label" in attributes, f"'Data label' row missing. Found: {attributes}"


def test_output_has_power_rail_attributes(tmp_path: Path):
    """P_SOC and P_VCCCORE power rails must appear as attribute rows in the output."""
    _, xlsx = _run_parseall(tmp_path, FIXTURE_PARSEALL_INPUT)
    attributes = _load_attribute_column(xlsx)
    assert "P_SOC" in attributes, f"'P_SOC' attribute missing. Found: {attributes}"
    assert "P_VCCCORE" in attributes, f"'P_VCCCORE' attribute missing. Found: {attributes}"


def test_multiple_datasets_produce_multiple_columns(tmp_path: Path):
    """With 3 sub-folders, the Excel must contain at least 3 data columns (+ Attribute)."""
    _, xlsx = _run_parseall(tmp_path, FIXTURE_PARSEALL_INPUT)
    wb = openpyxl.load_workbook(str(xlsx), read_only=True)
    ws = wb.active
    max_col = ws.max_column
    wb.close()
    # column 1 = Attribute, columns 2+ = datasets
    assert max_col >= 4, f"Expected ≥4 columns (Attribute + 3 datasets), got {max_col}"


def test_ms_model_throughput_attribute(tmp_path: Path):
    """When an MS AI model output file is present, 'throughput (FPS)' should appear."""
    _, xlsx = _run_parseall(tmp_path, FIXTURE_PARSEALL_INPUT)
    attributes = _load_attribute_column(xlsx)
    assert "throughput (FPS)" in attributes, (
        f"'throughput (FPS)' attribute missing. Found: {attributes}"
    )


def test_llama_throughput_attribute(tmp_path: Path):
    """When a Llama BM output file is present, '2nd tokens throughput (tok/s)' should appear."""
    _, xlsx = _run_parseall(tmp_path, FIXTURE_PARSEALL_INPUT)
    attributes = _load_attribute_column(xlsx)
    assert "2nd tokens throughput (tok/s)" in attributes, (
        f"Llama throughput attribute missing. Found: {attributes}"
    )

