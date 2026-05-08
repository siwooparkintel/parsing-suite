import pandas as pd
import parsers.tools as tools
import parsers.flattener as flattener 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def addKeyAutoHide(summary_type, flattened_socwatch_list):
    for idx, item in enumerate(flattened_socwatch_list) :
        if summary_type == "compact" and idx != 0 :
            item["auto-hide"] = True
        else :
            item["auto-hide"] = False

def autoHideColumn(excel_path):
    # detecting and hide cell named "auto-hide"

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    sheetA = sheet["A"]
    auto_hide_idx = None
    auto_hide_row = None
    
    for item in sheetA :
        # print(f"row : {item.row}, and value : {item.value}")
        if item.value == "auto-hide" :
            auto_hide_idx = item.row
            break

    if auto_hide_idx is not None and auto_hide_idx <= sheet.max_row :
        auto_hide_row = sheet[auto_hide_idx]

    for cell in auto_hide_row :
        if cell.value == True:
            sheet.column_dimensions[cell.column_letter].hidden = cell.value

    workbook.save(excel_path)
    # print(auto_hide_row)

def flatten_data_with_autohide(entry, picks, socwatch_targets, PCIe_targets):
    flatten_list = list()
    flattened = {'Data_label': entry['data_label'], 'Condition': entry['condition']}
    flattened.update(flattener.flatten_power_dic(entry, picks))
    flattened_socwatch_list = flattener.flatten_socwatch_dic_per_core(entry, socwatch_targets)
    
    addKeyAutoHide(entry['data_summary_type'], flattened_socwatch_list)

    flattened.update(flattened_socwatch_list[0])
    flattened.update(flattener.flatten_pcie_socwatch_dic(entry, PCIe_targets))
    flatten_list.append(flattened)
    flatten_list.extend(flattened_socwatch_list[1:]) if len(flattened_socwatch_list) > 1 else None
    return flatten_list

def flatten_data(entry, socwatch_targets, PCIe_targets, picks):
    flattened = {'Data label': entry['data_label'][0], 'Condition': entry['data_label'][1]}
    flattened.update(flattener.flatten_power_dic(entry, picks))
    flattened.update({"Data Detected": entry["data_type"]}) if "data_type" in entry else None
    flattened.update(flattener.flatten_ETL_dic(entry))
    flattened.update(flattener.flatten_AI_model_dic(entry))
    flattened.update(flattener.flatten_fps_dic(entry))
    flattened.update(flattener.flatten_lpmode_full_dic(entry))
    flattened.update(flattener.flatten_LPmode_sr_dic(entry))
    flattened.update(entry["procyon_score_obj"]) if "procyon_score_obj" in entry else None
    flattened.update(flattener.flatten_teams_vpt_camera_dic(entry))
    flattened.update(flattener.flatten_socwatch_dic(entry, socwatch_targets))
    flattened.update(flattener.flatten_pcie_socwatch_dic(entry, PCIe_targets))
    flattened.update(flattener.flatten_procyon_arielle_dic(entry))
    return flattened

def create_V_H_Excel(df, result_path):
    #df.to_excel(result_path+"_allPower_h.xlsx", index=False)
    df_v = df.transpose()
    df_v = df_v.reset_index()
    df_v.rename(columns={'index': 'Attribute'}, inplace=True)
    df_v.to_excel(result_path+"_allPower_v.xlsx", index=False)
    print(f"Excel files created at {result_path}_allPower_h.xlsx and {result_path}_allPower_v.xlsx")

def reportAllPowerAndType(result_path, hobl_data, socwatch_targets, PCIe_targets, picks) :
    #new_df_columns = flattener.getHeaderCollection(hobl_data, picks)
    flatten_data_list = list()
    for entry in hobl_data :
        if len(entry['data_type']) > 0 :
            flatten_data_list.append(flatten_data(entry, socwatch_targets, PCIe_targets, picks))
    
    # flatten_data_list = [flatten_data(entry, socwatch_targets, PCIe_targets, picks) if len(entry['data_type']) > 0 else None for entry in hobl_data]
    df = pd.DataFrame(flatten_data_list, columns=flattener.getHeaderCollection())
    create_V_H_Excel(df, result_path)




# def flatten_mlc_data(entry, socwatch_targets, PCIe_targets, picks):
#     flattened = {'Data label': entry['data_label'][0], 'Condition': entry['data_label'][1]}
#     flattened.update(flattener.flatten_power_dic(entry, picks))
#     flattened.update(flattener.flatten_mlc_output_dic(entry))
#     flattened.update(flattener.flatten_socwatch_dic(entry, socwatch_targets))
#     flattened.update(flattener.flatten_pcie_socwatch_dic(entry, PCIe_targets))
#     return flattened

# def reportAllPowerAndMLC(result_path, hobl_data, socwatch_targets, PCIe_targets, picks) :
#     df = pd.DataFrame([flatten_mlc_data(entry, socwatch_targets, PCIe_targets, picks) for entry in hobl_data])
#     create_V_H_Excel(df, result_path)

                



